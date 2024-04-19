import sys
import datetime
import zoneinfo
import openpyxl


def print_values_in_range(sheetdata, sheetmath, range_string, out=sys.stdout):
    """Print values of cells in a range.

    Print cell-values even if the cell is a formula.
    セルが数式であってもセルの値を表示します。

    Args:
        sheetdata (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to True
        sheetmath (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to False
        range_string (str) : Excel style cell range.
            example: `A1:D3`
        out (io): instance of I/O
            default: stdout

    Returns:
        None
    """
    (min_col, min_row, max_col,
     max_row) = openpyxl.utils.cell.range_boundaries(range_string)
    for row in sheetdata.iter_rows(min_row=min_row, max_row=max_row,
                                   min_col=min_col, max_col=max_col, values_only=True):
        for cell1 in row:
            out.write(str(cell1))
            out.write("\n")
    return


def print_formulas_in_range(sheetdata, sheetmath, range_string, out=sys.stdout):
    """Print formulas of cells in a range.

    If the cell is formula, print the formula.
    Otherwire print the cell-value.
    セルのデータが数式なら数式を表示します。そうでなければ値を表示します。

    Args:
        sheetdata (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to True
        sheetmath (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to False
        range_string (str): Excel style cell range.
            example: `A1:D3`
        out (io): instance of I/O
            default: stdout

    Returns:
        None
    """
    (min_col, min_row, max_col,
     max_row) = openpyxl.utils.cell.range_boundaries(range_string)
    for row in sheetmath.iter_rows(min_row=min_row, max_row=max_row,
                                   min_col=min_col, max_col=max_col, values_only=True):
        for cell1 in row:
            # print(cell1)
            out.write(str(cell1))
            out.write("\n")
    return


def get_creator_lastmodify(workbook):
    """Returns the creator and lastmodifiedby.

    Returns the creator and lastmodifiedby of the book.
    ファイルの作成者、最終更新者を返します。

    Args:
        workbook (WorkBook): WorkBook instance of the openpyxl

    Returns:
        (creator, lastmodifiedby) : tuple of strs
            (作成者, 最終更新者)
    """
    return (workbook.properties.creator, workbook.properties.lastModifiedBy)



def get_createtime_modifiedtime(workbook, iana_key='Asia/Tokyo'):
    """Returns the createdtime and lastmodifiedtime.

    Returns the created datetime and the lastmodified datetime of the book.
    The default timezone info is JST.
    ファイルの作成日時、最終更新日時を返します。
    デフォルトのタイムゾーンは日本標準時間です。

    Args:
        workbook (WorkBook): WorkBook instance of the openpyxl

        iana_key (str | optional): IANA timezone identifier
            default: 'Asia/Tokyo'

    Returns:
        (createdtime, lastmodifiedtime) : tuple of strs
            (作成者, 最終更新者)
            The datatimes are isoformat strings.
    """
    # # get datetime with "Z", (UTC)
    createdtime = workbook.properties.created
    modifiedtime = workbook.properties.modified
    # ただし、時間帯情報　timezone はNULLである　つまりシステム依存の時間に見えてしまう。
    # 日本時間に変換
    # 強引にUTCと認識させ、そこから日本時間帯に変換させる
    tmp = createdtime.replace(tzinfo=datetime.timezone.utc)
    createdtimeJST = tmp.astimezone(tz=zoneinfo.ZoneInfo(key=iana_key))
    tmp = modifiedtime.replace(tzinfo=datetime.timezone.utc)
    modifiedtimeJST = tmp.astimezone(tz=zoneinfo.ZoneInfo(key=iana_key))
    return (createdtimeJST, modifiedtimeJST)


def is_given_value(sheetdata, sheetmath, addr, value):
    """Verifies that the value of the specified cell is the given value.

    Returns True if the cell-value at the address is the given value.
    アドレスを指定されたセルの値が与えられた値であれば True を返します。

    Args:
        sheetdata (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to True
        sheetmath (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to False
        addr (str): Excel style cell address.
            example: "A3"
        value (int | str): value

    Returns:
        bool
    """
    if sheetdata[addr].value == value:
        return True
    else:
        return False


def is_formula(sheetdata, sheetmath, addr):
    """Verifies that the the specified cell is a formula.

    Returns True if the cell at the address is a formula.
    This function does not check the detail of the formula.
    アドレスを指定されたセルが数式であれば True を返します。
    数式の詳細には触れず、数式であれば True を返します。

    Args:
        sheetdata (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to True
        sheetmath (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to False
        addr (str): Excel style cell address.
            example: "A3"

    Returns:
        bool
    """
    if sheetdata[addr].value != sheetmath[addr].value:
        # 保存されている値とセルに書かれているデータが違う　→　セルに書かれているのは数式
        return True
    else:
        return False


def check_values_in_range(sheetdata, sheetmath, range_string, values):
    """Verifies that the cells in the range are the given values.

    Returns total number of cells in the range
    and number of cells whose values are equal to the given values.
    範囲を指定されたセルの数と、与えられた値と同じ値を持つセルの数を返します。

    Args:
        sheetdata (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to True
        sheetmath (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to False
        range_string (str): Excel style cell range.
            example: `A1:B3`
        values (list of list of int or str): Two dimensional values
            Row-major two dimensional sequence.
            example: [[11,12],[21,22],[31,32]]

    Returns:
        (countCells, countTrue) : tuple of ints
            countCells (int): total number of cells in the range
            countTrue (int): number of cells whose values are equal to the given values
    """
    (min_col, min_row, max_col,
     max_row) = openpyxl.utils.cell.range_boundaries(range_string)
    countCells = 0
    countTrue = 0
    for row, valrow in zip(sheetdata.iter_rows(min_row=min_row, max_row=max_row,
                                               min_col=min_col, max_col=max_col, values_only=True),
                           values):
        for cell1, val1 in zip(row, valrow):
            countCells += 1
            if cell1 == val1:
                countTrue += 1
    return (countCells, countTrue)


def check_values_in_range_float(sheetdata, sheetmath, range_string, values, diffval=0.01):
    """Verifies that the cells in the range are the given float values.

    Returns total number of cells in the range
    and number of cells whose values are considered to be the same as the given float values.
    Slight differences in values are acceptable.
    範囲を指定されたセルの数と、与えられた値と同じ値を持つと考えられるセルの数を返します。
    わずかな値の差は許容される。

    Args:
        sheetdata (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to True
        sheetmath (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to False
        range_string (str): Excel style cell range.
            example: `A1:B3`
        values (list of list of int or str): Two dimensional values
            Row-major two dimensional sequence.
            example: [[11.1,12.1],[21.1,22.1],[31.1,32.1]]
        diffval (float): Acceptable deviation
            10.12 and 10.123 are considerd to be the same if diffval is 0.01 .

    Returns:
        (countCells, countTrue) : tuple of ints
            countCells (int): total number of cells in the range
            countTrue (int): number of cells whose values are 'equal' to the given values
    """
    (min_col, min_row, max_col,
     max_row) = openpyxl.utils.cell.range_boundaries(range_string)
    countCells = 0
    countTrue = 0
    for row, valrow in zip(sheetdata.iter_rows(min_row=min_row, max_row=max_row,
                                               min_col=min_col, max_col=max_col, values_only=True),
                           values):
        for cell1, val1 in zip(row, valrow):
            countCells += 1
            if cell1 is not None and \
                (isinstance(cell1, int) or isinstance(cell1, float)) and \
                    abs(cell1 - val1) <= diffval:
                countTrue += 1
    return (countCells, countTrue)


def check_num_formulas_in_range(sheetdata, sheetmath, range_string):
    """Verifies that the cells in the range are formulas.

    Returns total number of cells in the range
    and number of cells whose data are formulas.
    This function does not check the detail of the formula.
    範囲を指定されたセルの数と、数式であるセルの数を返します。
    数式の詳細には触れず、数式であればカウントします。

    Args:
        sheetdata (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to True
        sheetmath (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to False
        range_string (str): Excel style cell range.
            example: `A1:B3`

    Returns:
        (countCells, countTrue) : tuple of ints
            countCells (int): total number of cells in the range
            countTrue (int): number of cells which are formulas
    """
    (min_col, min_row, max_col,
     max_row) = openpyxl.utils.cell.range_boundaries(range_string)
    countCells = 0
    countTrue = 0
    for rowdata, rowmath in zip(sheetdata.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col, values_only=True),
                                sheetmath.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col, values_only=True)):
        for data1, math1 in zip(rowdata, rowmath):
            countCells += 1
            if data1 != math1:
                # 保存されている値とセルに書かれているデータが違う　→　セルに書かれているのは数式
                countTrue += 1
    return (countCells, countTrue)


def check_func_in_range(sheetdata, sheetmath, range_string, func_string):
    """Verifies that the formulas contain the given function.

    Returns total number of cells in the range
    and number of cells whose formula contains the given function.
    This function checks the formula and function as string.
    So this function do not distinguish SUM and SUMIF and so on.
    範囲を指定されたセルの数と、与えられた関数を含む数式のセルの数を返します。
    数式の詳細には触れず、文字列として関数名を含むかどうかをチェックします。
    その為、SUM と SUMIF などは区別しません。

    Args:
        sheetdata (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to True
        sheetmath (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to False
        range_string (str): Excel style cell range.
            example: `A1:B3`
        func_string (str): Excel function name
            example : 'AVERAGE'

    Returns:
        (countCells, countTrue) : tuple of ints
            countCells (int): total number of cells in the range
            countTrue (int): number of cells which contains the given function
    """
    (min_col, min_row, max_col,
     max_row) = openpyxl.utils.cell.range_boundaries(range_string)
    countCells = 0
    countTrue = 0
    for row in sheetmath.iter_rows(min_row=min_row, max_row=max_row,
                                   min_col=min_col, max_col=max_col, values_only=True):
        countCells += 1
        for cell1 in row:
            if (isinstance(cell1, str)) and \
                    func_string in cell1:
                countTrue += 1
    return (countCells, countTrue)


def is_integer(sheetdata, sheetmath, addr):
    """Verifies that the value of the specified cell is int.

    Returns True if the cell-value at the address is integer.
    アドレスを指定されたセルの値が整数型であれば True を返します。

    Args:
        sheetdata (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to True
        sheetmath (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to False
        addr (str): Excel style cell address.
            example: "A3"

    Returns:
        bool
    """
    if isinstance(sheetdata[addr].value, int):
        return True
    else:
        return False


def check_comp_abs_ref_in_range(sheetdata, sheetmath, range_string):
    """Verifies that the cells in the range contains composite or absolute cell reference.

    Returns total number of cells in the range
    and number of cells whose formula contains composite/absolute cell reference.
    範囲を指定されたセルの数と、セルの複合参照もしくは絶対参照を含む数式のセルの数を返します。

    Args:
        sheetdata (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to True
        sheetmath (Worksheet): Worksheet instance of the openpyxl
            whose book are opened with data_only set to False
        addr (str): Excel style cell address.
            example: "A3"

    Returns:
        (countCells, countTrue) : tuple of ints
            countCells (int): total number of cells in the range
            countTrue (int): number of cells which contains composite/absolute cell reference
    """
    return check_func_in_range(sheetdata, sheetmath, range_string, func_string="$")


# 指定したセルの「フォント」がｘｘである
# 作成せず

# 指定したセルの「罫線」がｘｘである
# 作成せず

# 指定したセル範囲の「条件付き書式」がある
# 作成せず

# グラフ
# 作成せず
